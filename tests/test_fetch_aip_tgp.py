"""Hermetic tests for fetch.aip_tgp.

Both the index page (HTML) and the weekly XLSX download are mocked.
The XLSX is constructed in-memory with the real layout the AIP file has.
"""
from __future__ import annotations

import datetime as dt
import io
import time
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
import responses

from fuel_pred.fetch import aip_tgp

# ----------------------------- Fixtures -----------------------------


def _build_aip_xlsx(petrol_rows: list[tuple[dt.date, float]],
                    diesel_rows: list[tuple[dt.date, float]]) -> bytes:
    wb = openpyxl.Workbook()
    notes = wb.active
    assert notes is not None
    notes.title = "NOTES"
    notes["B7"] = "TERMINAL GATE PRICES (TGPs)"

    petrol = wb.create_sheet(aip_tgp.PETROL_SHEET)
    diesel = wb.create_sheet(aip_tgp.DIESEL_SHEET)

    headers = [
        "AVERAGE ULP TGPS\n(inclusive of GST)",
        "Sydney", "Melbourne", "Brisbane", "Adelaide",
        "Perth", "Darwin", "Hobart", "National\nAverage",
    ]
    for col, h in enumerate(headers, start=1):
        petrol.cell(row=1, column=col, value=h)
        diesel.cell(row=1, column=col, value=h)

    for r, (date, value) in enumerate(petrol_rows, start=2):
        petrol.cell(row=r, column=1, value=date)
        petrol.cell(row=r, column=2, value=value)  # Sydney
        petrol.cell(row=r, column=9, value=value + 1)  # National avg (irrelevant)
    for r, (date, value) in enumerate(diesel_rows, start=2):
        diesel.cell(row=r, column=1, value=date)
        diesel.cell(row=r, column=2, value=value)
        diesel.cell(row=r, column=9, value=value + 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def index_html() -> bytes:
    """Synthetic AIP index page with two dated XLSX links."""
    return (
        b"<html><body>"
        b'<a href="/sites/default/files/download-files/2024-04/'
        b'AIP_TGP_Data_26-Apr-2024.xlsx">Older</a>'
        b'<a href="/sites/default/files/download-files/2024-05/'
        b'AIP_TGP_Data_03-May-2024.xlsx">Latest</a>'
        b'<a href="/somewhere/else.pdf">Distractor</a>'
        b"</body></html>"
    )


@pytest.fixture
def xlsx_payload() -> bytes:
    return _build_aip_xlsx(
        petrol_rows=[
            (dt.date(2024, 1, 1), 175.5),
            (dt.date(2024, 1, 2), 176.0),
            (dt.date(2024, 1, 3), 176.5),
        ],
        diesel_rows=[
            (dt.date(2024, 1, 1), 195.0),
            (dt.date(2024, 1, 2), 195.5),
            (dt.date(2024, 1, 3), 196.0),
        ],
    )


# ----------------------------- discover_latest_xlsx_url -----------------------------


def test_discover_latest_xlsx_url_picks_lexicographically_last(index_html: bytes) -> None:
    url = aip_tgp.discover_latest_xlsx_url(index_html)
    assert url.endswith("AIP_TGP_Data_03-May-2024.xlsx")
    assert "2024-05" in url


def test_discover_latest_xlsx_url_raises_when_no_match() -> None:
    html = b"<html><body><a href='/some/other/file.pdf'>X</a></body></html>"
    with pytest.raises(RuntimeError, match="no AIP_TGP_Data"):
        aip_tgp.discover_latest_xlsx_url(html)


# ----------------------------- parse_aip_xlsx -----------------------------


def test_parse_aip_xlsx_merges_petrol_and_diesel(xlsx_payload: bytes) -> None:
    df = aip_tgp.parse_aip_xlsx(xlsx_payload)
    assert list(df.columns) == ["date", "ulp_sydney", "diesel_sydney"]
    assert len(df) == 3
    jan1 = df["date"] == dt.date(2024, 1, 1)
    assert df.loc[jan1, "ulp_sydney"].iloc[0] == pytest.approx(175.5)
    assert df.loc[jan1, "diesel_sydney"].iloc[0] == pytest.approx(195.0)


def test_parse_aip_xlsx_raises_on_missing_sheet() -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "Wrong"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(RuntimeError, match=aip_tgp.PETROL_SHEET):
        aip_tgp.parse_aip_xlsx(buf.getvalue())


def test_parse_aip_xlsx_raises_when_sydney_column_missing() -> None:
    wb = openpyxl.Workbook()
    wb.active.title = aip_tgp.PETROL_SHEET
    wb.active.cell(row=1, column=1, value="DATE")
    wb.active.cell(row=1, column=2, value="Melbourne")  # no Sydney
    diesel = wb.create_sheet(aip_tgp.DIESEL_SHEET)
    diesel.cell(row=1, column=1, value="DATE")
    diesel.cell(row=1, column=2, value="Sydney")
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(RuntimeError, match=aip_tgp.SYDNEY_COLUMN):
        aip_tgp.parse_aip_xlsx(buf.getvalue())


# ----------------------------- fetch (full path) -----------------------------


@responses.activate
def test_fetch_writes_parquet_with_expected_schema(
    tmp_path: Path, index_html: bytes, xlsx_payload: bytes
) -> None:
    out = tmp_path / "aip_tgp.parquet"
    responses.add(responses.GET, aip_tgp.INDEX_URL, body=index_html, status=200)
    responses.add(
        responses.GET,
        aip_tgp.BASE_URL
        + "/sites/default/files/download-files/2024-05/AIP_TGP_Data_03-May-2024.xlsx",
        body=xlsx_payload,
        status=200,
    )

    aip_tgp.fetch("2024-01-01", "2024-12-31", out)

    df = pd.read_parquet(out)
    assert list(df.columns) == ["date", "ulp_sydney", "diesel_sydney"]
    assert len(df) == 3


@responses.activate
def test_fetch_filters_to_requested_range(
    tmp_path: Path, index_html: bytes
) -> None:
    out = tmp_path / "aip_tgp.parquet"
    payload = _build_aip_xlsx(
        petrol_rows=[
            (dt.date(2023, 12, 31), 170.0),
            (dt.date(2024, 1, 1), 175.5),
            (dt.date(2024, 1, 31), 180.0),
            (dt.date(2024, 12, 31), 200.0),
        ],
        diesel_rows=[
            (dt.date(2024, 1, 1), 195.0),
            (dt.date(2024, 1, 31), 198.0),
        ],
    )
    responses.add(responses.GET, aip_tgp.INDEX_URL, body=index_html, status=200)
    responses.add(
        responses.GET,
        aip_tgp.BASE_URL
        + "/sites/default/files/download-files/2024-05/AIP_TGP_Data_03-May-2024.xlsx",
        body=payload,
        status=200,
    )

    aip_tgp.fetch("2024-01-01", "2024-06-30", out)
    df = pd.read_parquet(out)
    # The merge is outer on date, then filter to range. Both Jan dates
    # have ULP; only Jan dates have Diesel. Range covers Jan 1..Jun 30.
    assert df["date"].min().isoformat() == "2024-01-01"
    assert df["date"].max().isoformat() == "2024-01-31"


@responses.activate
def test_fetch_skips_when_cache_fresh(tmp_path: Path) -> None:
    out = tmp_path / "aip_tgp.parquet"
    out.write_bytes(b"placeholder")
    aip_tgp.fetch("2024-01-01", "2024-12-31", out, max_age_days=7.0)
    assert out.read_bytes() == b"placeholder"


@responses.activate
def test_fetch_force_bypasses_cache(
    tmp_path: Path, index_html: bytes, xlsx_payload: bytes
) -> None:
    out = tmp_path / "aip_tgp.parquet"
    out.write_bytes(b"placeholder")
    responses.add(responses.GET, aip_tgp.INDEX_URL, body=index_html, status=200)
    responses.add(
        responses.GET,
        aip_tgp.BASE_URL
        + "/sites/default/files/download-files/2024-05/AIP_TGP_Data_03-May-2024.xlsx",
        body=xlsx_payload,
        status=200,
    )
    aip_tgp.fetch("2024-01-01", "2024-12-31", out, force=True)
    df = pd.read_parquet(out)
    assert len(df) == 3


@responses.activate
def test_fetch_stale_cache_triggers_refetch(
    tmp_path: Path, index_html: bytes, xlsx_payload: bytes
) -> None:
    import os

    out = tmp_path / "aip_tgp.parquet"
    out.write_bytes(b"placeholder")
    old = time.time() - 14 * 86400
    os.utime(out, (old, old))
    responses.add(responses.GET, aip_tgp.INDEX_URL, body=index_html, status=200)
    responses.add(
        responses.GET,
        aip_tgp.BASE_URL
        + "/sites/default/files/download-files/2024-05/AIP_TGP_Data_03-May-2024.xlsx",
        body=xlsx_payload,
        status=200,
    )
    aip_tgp.fetch("2024-01-01", "2024-12-31", out, max_age_days=7.0)
    df = pd.read_parquet(out)
    assert len(df) == 3


@responses.activate
def test_fetch_empty_range_raises(
    tmp_path: Path, index_html: bytes, xlsx_payload: bytes
) -> None:
    out = tmp_path / "aip_tgp.parquet"
    responses.add(responses.GET, aip_tgp.INDEX_URL, body=index_html, status=200)
    responses.add(
        responses.GET,
        aip_tgp.BASE_URL
        + "/sites/default/files/download-files/2024-05/AIP_TGP_Data_03-May-2024.xlsx",
        body=xlsx_payload,
        status=200,
    )
    with pytest.raises(RuntimeError, match="no AIP TGP"):
        aip_tgp.fetch("2030-01-01", "2030-12-31", out)
