"""Fetch AIP Sydney Terminal Gate Prices for ULP and Diesel.

Source: https://www.aip.com.au/historical-ulp-and-diesel-tgp-data
        (file URL discovered by parsing the index page; pattern is
        ``<base>/sites/default/files/download-files/<YYYY-MM>/AIP_TGP_Data_<DD-MMM-YYYY>.xlsx``)

The "weekly" XLSX is misnamed — it actually contains the **full daily
TGP history back to 2004-01-01** for all 7 capital cities + national
average. Two sheets: ``Petrol TGP`` (ULP) and ``Diesel TGP``. Each
release adds the new days since the previous publication; older rows
are unchanged.

The spec hint in §13 Q1 anticipated that AIP would be forward-only
("scrape forward from project start; partial backfill via Wayback if
motivated"). Reality is the file gives us the full history in one
download. The `upstream_tgp_*` features in §7.2 can therefore be fully
populated for the entire 2016-09 → present span.

Granularity: daily, business days only.
Coverage: 2004-01-01 → most recent week.

Output schema: ``date, ulp_sydney, diesel_sydney`` (cents per litre,
inclusive of GST). Other cities + national average are dropped — this
project is NSW only and the spec only references Sydney TGP (§7.2).

Spec: spec.md §5.2, §7.2, §13 Q1.
"""
from __future__ import annotations

import argparse
import io
import logging
import re
import time
from pathlib import Path

import openpyxl
import pandas as pd

from fuel_pred.fetch._ckan import download_bytes

logger = logging.getLogger(__name__)

INDEX_URL: str = "https://www.aip.com.au/historical-ulp-and-diesel-tgp-data"
BASE_URL: str = "https://www.aip.com.au"

# Pattern of the dated weekly XLSX inside the index page.
_WEEKLY_XLSX_RE: re.Pattern[str] = re.compile(
    r"/sites/default/files/download-files/\d{4}-\d{2}/AIP_TGP_Data_[^\"' ]+?\.xlsx",
    re.IGNORECASE,
)

# Sheet names in the weekly file (verified May 2026).
PETROL_SHEET: str = "Petrol TGP"
DIESEL_SHEET: str = "Diesel TGP"

# Header column we extract from each sheet.
SYDNEY_COLUMN: str = "Sydney"


def _is_cache_fresh(out: Path, max_age_days: float) -> bool:
    if not out.exists():
        return False
    age_days = (time.time() - out.stat().st_mtime) / 86400.0
    return age_days < max_age_days


def discover_latest_xlsx_url(index_html: bytes) -> str:
    """Find the URL of the most recent weekly TGP XLSX in the index HTML.

    Strategy: regex-extract every `/AIP_TGP_Data_*.xlsx` href, sort
    lexicographically (the date-stamped path naturally sorts to latest
    last because the date encoding is `<YYYY-MM>/AIP_TGP_Data_<DD-MMM-YYYY>`,
    and the year-month subdirectory dominates), and return the last one.

    Raises if no match.
    """
    text = index_html.decode("utf-8", errors="replace")
    matches: list[str] = sorted(set(_WEEKLY_XLSX_RE.findall(text)))
    if not matches:
        raise RuntimeError(
            "no AIP_TGP_Data_*.xlsx links found on AIP index page; "
            "the page layout may have changed"
        )
    return BASE_URL + matches[-1]


def parse_aip_xlsx(payload: bytes) -> pd.DataFrame:
    """Parse the weekly XLSX into a (date, ulp_sydney, diesel_sydney) frame.

    Public so tests can call it with synthetic XLSX bytes.
    """
    wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    if PETROL_SHEET not in wb.sheetnames or DIESEL_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"AIP XLSX missing expected sheets {PETROL_SHEET!r} / {DIESEL_SHEET!r}; "
            f"have: {wb.sheetnames}"
        )

    petrol = _read_sheet(wb[PETROL_SHEET], "ulp_sydney")
    diesel = _read_sheet(wb[DIESEL_SHEET], "diesel_sydney")
    return petrol.merge(diesel, on="date", how="outer").sort_values("date").reset_index(drop=True)


def _read_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, value_col: str) -> pd.DataFrame:
    """Read one TGP sheet's date column + Sydney column into a DataFrame."""
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    if SYDNEY_COLUMN not in header:
        raise RuntimeError(
            f"AIP sheet header missing {SYDNEY_COLUMN!r} column; have: {header}"
        )
    sydney_idx = header.index(SYDNEY_COLUMN)

    dates: list[object] = []
    values: list[object] = []
    for row in rows:
        date_cell = row[0]
        value_cell = row[sydney_idx]
        if date_cell is None:
            continue
        dates.append(date_cell)
        values.append(value_cell)

    df = pd.DataFrame({"date": dates, value_col: values})
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
    df[value_col] = pd.to_numeric(df[value_col], errors="coerce")
    df = df[df["date"].notna() & df[value_col].notna()].reset_index(drop=True)
    return df


def fetch(
    start: str,
    end: str,
    out: Path,
    *,
    force: bool = False,
    max_age_days: float = 7.0,
) -> None:
    """Fetch AIP Sydney TGP for ULP + Diesel and write Parquet
    ``date, ulp_sydney, diesel_sydney``.

    Args:
        start: ISO date, inclusive.
        end: ISO date, inclusive.
        out: output Parquet path.
        force: re-fetch ignoring cache.
        max_age_days: skip re-fetch when cache is fresher than this.
            Default 7 days — AIP publishes weekly.
    """
    if not force and _is_cache_fresh(out, max_age_days):
        logger.info("cache hit %s (< %.0f days old) — skipping fetch", out, max_age_days)
        return

    logger.info("looking up latest AIP weekly XLSX from %s", INDEX_URL)
    index_html = download_bytes(INDEX_URL)
    xlsx_url = discover_latest_xlsx_url(index_html)
    logger.info("downloading %s", xlsx_url)
    payload = download_bytes(xlsx_url)
    df = parse_aip_xlsx(payload)

    start_date = pd.to_datetime(start).date()
    end_date = pd.to_datetime(end).date()
    df = df[(df["date"] >= start_date) & (df["date"] <= end_date)].copy()

    if df.empty:
        raise RuntimeError(f"no AIP TGP rows in range {start}..{end}")

    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(out, engine="pyarrow", compression="zstd", index=False)
    logger.info("wrote %d rows to %s", len(df), out)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--start", required=True)
    parser.add_argument("--end", required=True)
    parser.add_argument("--out", required=True, type=Path)
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--max-age-days", type=float, default=7.0)
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(name)s %(levelname)s %(message)s")
    fetch(args.start, args.end, args.out, force=args.force, max_age_days=args.max_age_days)


if __name__ == "__main__":
    main()
