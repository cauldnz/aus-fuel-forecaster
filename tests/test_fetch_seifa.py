"""Hermetic tests for fetch.seifa. ABS xlsx mocked with `responses`."""
from __future__ import annotations

import io
import time
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
import responses

from fuel_pred.fetch import seifa


def _build_seifa_xlsx(rows: list[tuple[object, ...]]) -> bytes:
    """Build a synthetic SEIFA xlsx with the real layout (Table 1, header row 6)."""
    wb = openpyxl.Workbook()
    # ABS file has Contents + Tables 1-6 + Explanatory Notes; we only need Table 1.
    default_sheet = wb.active
    assert default_sheet is not None
    default_sheet.title = "Contents"
    ws = wb.create_sheet(seifa.SHEET)

    # Real preamble: rows 1-5 have title / release / sub-headers.
    ws["A1"] = "Australian Bureau of Statistics"
    ws["A2"] = "Socio-Economic Indexes for Australia (SEIFA), 2021"
    ws["A4"] = "Table 1 Statistical Area Level 2 (SA2) SEIFA Summary, 2021"
    # Row 5: index group header (just a label band)
    ws.cell(row=5, column=3, value="Index of Relative Socio-economic Disadvantage")
    # Row 6: column headers
    headers = [
        "2021 Statistical Area Level 2 (SA2) 9-Digit Code",
        "2021 Statistical Area Level 2 (SA2) Name",
        "Score",
        "Decile",
        "Score",
        "Decile",
        "Score",
        "Decile",
        "Score",
        "Decile",
        "Usual Resident Population",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=seifa.HEADER_ROW, column=col, value=h)
    # Row 7+: data
    for r, row in enumerate(rows, start=seifa.HEADER_ROW + 1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


SAMPLE_ROWS: list[tuple[object, ...]] = [
    # (sa2_code, sa2_name, irsd_s, irsd_d, irsad_s, irsad_d, ier_s, ier_d, ieo_s, ieo_d, urp)
    (101021007, "Braidwood", 1024, 6, 1001, 6, 1027, 7, 1008, 6, 4343),
    (117011635, "Mascot", 1098, 9, 1110, 10, 1095, 9, 1124, 10, 21573),
    (122021422, "Newport - Bilgola", 1102, 10, 1118, 10, 1080, 8, 1141, 10, 13681),
    # A non-SA2 rollup row (state total) — must be dropped by the regex filter.
    ("New South Wales", "NSW", 1000, 5, 1000, 5, 1000, 5, 1000, 5, 8000000),
]


@responses.activate
def test_writes_parquet_with_expected_schema(tmp_path: Path) -> None:
    out = tmp_path / "seifa.parquet"
    responses.add(responses.GET, seifa.URL, body=_build_seifa_xlsx(SAMPLE_ROWS), status=200)

    seifa.fetch(out)

    df = pd.read_parquet(out)
    assert list(df.columns) == list(seifa.COLUMNS)
    # 3 valid SA2 rows; the "New South Wales" rollup row should be filtered.
    assert len(df) == 3
    assert "New South Wales" not in df["sa2_code"].tolist()


@responses.activate
def test_sa2_code_is_string_and_zero_padded_friendly(tmp_path: Path) -> None:
    out = tmp_path / "seifa.parquet"
    responses.add(responses.GET, seifa.URL, body=_build_seifa_xlsx(SAMPLE_ROWS), status=200)

    seifa.fetch(out)
    df = pd.read_parquet(out)
    # SA2 code must be string-shaped for join consistency with the
    # augmentor's output (string vs StringDtype both render as Python str).
    assert isinstance(df["sa2_code"].iloc[0], str)
    assert df["sa2_code"].iloc[0] == "101021007"


@responses.activate
def test_irsd_score_typed_as_int(tmp_path: Path) -> None:
    out = tmp_path / "seifa.parquet"
    responses.add(responses.GET, seifa.URL, body=_build_seifa_xlsx(SAMPLE_ROWS), status=200)

    seifa.fetch(out)
    df = pd.read_parquet(out)
    # IRSD scores cluster around 1000 (mean=1000, sd~100).
    assert (df["irsd_score"] > 800).all()
    assert (df["irsd_score"] < 1300).all()
    # Decile is 1-10.
    assert df["irsd_decile"].between(1, 10).all()


@responses.activate
def test_skips_when_cache_fresh(tmp_path: Path) -> None:
    out = tmp_path / "seifa.parquet"
    out.write_bytes(b"placeholder")

    # No mock registered; if the fetcher hits the network the test fails.
    seifa.fetch(out, max_age_days=1.0)

    assert out.read_bytes() == b"placeholder"


@responses.activate
def test_force_bypasses_cache(tmp_path: Path) -> None:
    out = tmp_path / "seifa.parquet"
    out.write_bytes(b"placeholder")
    responses.add(responses.GET, seifa.URL, body=_build_seifa_xlsx(SAMPLE_ROWS), status=200)

    seifa.fetch(out, force=True)
    df = pd.read_parquet(out)
    assert len(df) == 3


@responses.activate
def test_stale_cache_triggers_refetch(tmp_path: Path) -> None:
    import os

    out = tmp_path / "seifa.parquet"
    out.write_bytes(b"placeholder")
    old = time.time() - 800 * 86400  # > 365 days
    os.utime(out, (old, old))
    responses.add(responses.GET, seifa.URL, body=_build_seifa_xlsx(SAMPLE_ROWS), status=200)

    seifa.fetch(out, max_age_days=365.0)
    df = pd.read_parquet(out)
    assert len(df) == 3


def test_parser_raises_on_missing_sheet() -> None:
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(RuntimeError, match=seifa.SHEET):
        seifa.parse_seifa_xlsx(buf.getvalue())


def test_parser_raises_on_empty_table() -> None:
    """If the data section is empty, raise rather than silently writing 0 rows."""
    payload = _build_seifa_xlsx([])
    with pytest.raises(RuntimeError, match="no data rows"):
        seifa.parse_seifa_xlsx(payload)
