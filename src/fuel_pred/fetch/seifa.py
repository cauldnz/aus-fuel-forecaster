"""Fetch ABS SEIFA 2021 SA2-level scores.

Source: https://www.abs.gov.au/statistics/people/people-and-communities/socio-economic-indexes-areas-seifa-australia/latest-release
File:   `Statistical Area Level 2, Indexes, SEIFA 2021.xlsx` (Table 1)

The Excel file is a one-time fetch — SEIFA is Census-tied and won't be
updated until the 2026 SEIFA release in 2028. We still cache locally per
the project's caching philosophy: re-runs are free, content rarely
changes, no need for force-refresh logic.

Output schema (`data/raw/seifa_2021_sa2.parquet`):

| col | type | source |
|---|---|---|
| `sa2_code` | string | Table 1 col A (9-digit, kept as string for join consistency) |
| `sa2_name` | string | Table 1 col B |
| `irsd_score` | int | Table 1 col C — Index of Relative Socio-economic Disadvantage |
| `irsd_decile` | int | Table 1 col D |
| `irsad_score` | int | Table 1 col E — Index of Rel. Socio-economic Advantage and Disadvantage |
| `irsad_decile` | int | Table 1 col F |
| `ier_score` | int | Table 1 col G — Index of Economic Resources |
| `ier_decile` | int | Table 1 col H |
| `ieo_score` | int | Table 1 col I — Index of Education and Occupation |
| `ieo_decile` | int | Table 1 col J |
| `usual_resident_population` | int | Table 1 col K |

Spec: spec.md §5.4 (mentions SEIFA as joined separately from the augmentor).
"""
from __future__ import annotations

import argparse
import io
import logging
import time
from pathlib import Path

import openpyxl
import pandas as pd

from fuel_pred.fetch._ckan import download_bytes

logger = logging.getLogger(__name__)

URL: str = (
    "https://www.abs.gov.au/statistics/people/people-and-communities/"
    "socio-economic-indexes-areas-seifa-australia/2021/"
    "Statistical%20Area%20Level%202%2C%20Indexes%2C%20SEIFA%202021.xlsx"
)

# Verified May 2026: Table 1 sheet name + header row are stable across the
# 2021 release lifetime (no updates expected until 2028).
SHEET: str = "Table 1"
HEADER_ROW: int = 6  # 1-indexed; row 7+ holds data
COLUMNS: tuple[str, ...] = (
    "sa2_code",
    "sa2_name",
    "irsd_score",
    "irsd_decile",
    "irsad_score",
    "irsad_decile",
    "ier_score",
    "ier_decile",
    "ieo_score",
    "ieo_decile",
    "usual_resident_population",
)


def _is_cache_fresh(out: Path, max_age_days: float) -> bool:
    if not out.exists():
        return False
    age_days = (time.time() - out.stat().st_mtime) / 86400.0
    return age_days < max_age_days


def parse_seifa_xlsx(payload: bytes) -> pd.DataFrame:
    """Parse the SEIFA xlsx bytes into the canonical DataFrame.

    Public so tests can call it with synthetic xlsx payloads.
    """
    wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"sheet {SHEET!r} not found; available: {wb.sheetnames}")
    ws = wb[SHEET]

    rows: list[tuple[object, ...]] = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        # First cell is the SA2 9-digit code; skip blank trailers.
        if row[0] is None:
            continue
        rows.append(row[: len(COLUMNS)])

    if not rows:
        raise RuntimeError("SEIFA xlsx contained no data rows below the header")

    df = pd.DataFrame(rows, columns=list(COLUMNS))
    df["sa2_code"] = df["sa2_code"].astype(str).str.strip()
    df["sa2_name"] = df["sa2_name"].astype(str).str.strip()
    for col in COLUMNS[2:]:
        df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")

    # Filter out the "Australia" / state-rollup rows the workbook tucks at
    # the bottom — real SA2 codes are 9 digits.
    df = df[df["sa2_code"].str.fullmatch(r"\d{9}")].reset_index(drop=True)

    return df


def fetch(out: Path, *, force: bool = False, max_age_days: float = 365.0) -> None:
    """Download SEIFA xlsx and write parquet.

    Args:
        out: output Parquet path; parent dir created if missing.
        force: re-fetch even when cache is fresh.
        max_age_days: cache TTL. Default 365 days — SEIFA only updates
            on the Census cycle, ~5 years.
    """
    if not force and _is_cache_fresh(out, max_age_days):
        logger.info("cache hit %s (< %.0f days old) — skipping fetch", out, max_age_days)
        return

    payload = download_bytes(URL)
    df = parse_seifa_xlsx(payload)

    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(out, engine="pyarrow", compression="zstd", index=False)
    logger.info("wrote %d SA2 SEIFA rows to %s", len(df), out)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", required=True, type=Path)
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--max-age-days", type=float, default=365.0)
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(name)s %(levelname)s %(message)s")
    fetch(args.out, force=args.force, max_age_days=args.max_age_days)


if __name__ == "__main__":
    main()
